from  openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws1 = wb.create_sheet("mysheet")
ws2 = wb.create_sheet("mysheet", 0)
ws3 = wb.create_sheet("mysheet", -1)
ws.title ="New Title"
ws.sheet_properties.tabColor = "1072BA"


ws = wb["sample.xlsx"]




from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws["A1"] = 42
c = ws["A4"]
ws["A4"] = 4

d = ws.cell(row=4, column=2, value=10)
for x in range(1,10):
    for y in range(1,10):
        ws.cell(row=x, column=y)

cell_range = ws["A1" : "C2"]
colC = ws["C"]
col_range = ws["C:D"]
row10 = ws[10]
row_range = ws[5:10]
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)
for col in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in col:
        print(cell)

ws = wb.active
ws["C9"] = "my name is:"
print(tuple(ws.rows))

print(tuple(ws.columns))

for row in ws.values:
    for value in row:
        print(value)

c.value = "my name is: "
print(c.value)

d.value = 3.14
print(d.value)

# ws.append([1,2,3,4,5,6,7,8,9,10])

import datetime
ws["A1"] = datetime.datetime.now()

# save file to xlsx
wb = Workbook()
wb.save("balances.xlsx")







